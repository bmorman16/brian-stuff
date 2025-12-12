# run command: python grab_ticketsV3.py

import requests
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Replace these variables with your own
GITLAB_URL = 'https://git.web.boeing.com'  # or your GitLab instance URL
ACCESS_TOKEN = 'Zso1BU5ZFkyzBzuAC_ix'  # Your personal access token
PROJECT_ID = 129604# Your project ID


# Set up the headers for authentication
headers = {
    'Private-Token': ACCESS_TOKEN
}

# Function to get all issue_iid's and their comments from the GitLab project
def get_issues_and_comments():
    url = f'{GITLAB_URL}/api/v4/projects/{PROJECT_ID}/issues'
    issues_data = []
    page = 1

    while True:
        response = requests.get(url, headers=headers, params={'page': page, 'per_page': 100})
        if response.status_code != 200:
            print(f"Error fetching issues: {response.status_code} - {response.text}")
            break

        issues = response.json()
        if not issues:
            break  # No more issues to fetch

        for issue in issues:
            # Check if the issue is closed
            if issue['state'] == 'closed':
                continue  # Skip closed issues

            issue_iid = issue['iid']
            comments = get_comments(issue_iid)  # Fetch comments for the issue

   # Extract due date, time spent, and labels
            due_date = issue.get('due_date', 'N/A')  # Default to 'N/A' if not set
            time_spent = issue['time_stats'].get('total_time_spent', 0)  # Total time spent in seconds
            labels = issue.get('labels', [])  # Get labels, default to empty list if not set

            issues_data.append({
               'issue_iid': issue_iid,
                'title': issue['title'],
                'due_date': due_date,
                'time_spent': time_spent,
                'labels': ', '.join(labels),  # Join labels into a single string
                'comments': comments
            })
        page += 1  # Move to the next page
        #time.sleep(1)  # Optional: Add a delay to avoid hitting rate limits

    return issues_data

# Function to get comments for a specific issue
def get_comments(issue_iid):
    url = f'{GITLAB_URL}/api/v4/projects/{PROJECT_ID}/issues/{issue_iid}/notes'
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        comments = response.json()  # Return the list of comments
        # Filter out comments that start with "changed due date" or "changed title from"
        filtered_comments = [
            comment for comment in comments 
            if not (comment['body'].startswith("changed due date") or comment['body'].startswith("changed title from"))
        ]
        return filtered_comments
    else:
        print(f"Error fetching comments for issue {issue_iid}: {response.status_code} - {response.text}")
        return []

# Function to save issues and comments to an Excel file with formatting
def save_to_excel(issues_with_comments):
    # Prepare data for DataFrame
    data = []
    for issue in issues_with_comments:
        for comment in issue['comments']:
            data.append({
                'Issue IID': issue['issue_iid'],
                'Title': issue['title'],
                'Due Date': issue['due_date'],
                'Time Spent (seconds)': issue['time_spent'],
                'Labels': issue['labels'],
                'Comment ID': comment['id'],
                'Comment Body': comment['body'],
                'Comment Created At': comment['created_at']
            })

    # Create a DataFrame
    df = pd.DataFrame(data)

    # Define the Excel file name
    excel_file = 'gitlab_issues_with_comments.xlsx'
    
    # Save to Excel
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Issues')

        # Access the workbook and the worksheet
        workbook = writer.book
        worksheet = writer.sheets['Issues']

        # Set the header font to bold and add background color
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow background
        now_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green background
        next_fill = PatternFill(start_color='6699cc', end_color='6699cc', fill_type='solid')
        later_fill = PatternFill(start_color='9400d3', end_color='9400d3', fill_type='solid')
        spike_fill = PatternFill(start_color='dc143c', end_color='dc143c', fill_type='solid')
        
        for cell in worksheet[1]:  # The first row is the header
            cell.font = header_font
            cell.fill = header_fill

# Set column widths and apply light green fill for rows with label "Now"
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=len(df.columns)):
        # Set column widths
            for cell in row:
                max_length = 0
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[cell.column_letter].width = adjusted_width

        # Check if the label "Now" is in the current issue's labels
            if 'Now' in row[4].value:  # Assuming 'Labels' is the 5th column (index 4)
                for cell in row:
                    cell.fill = now_fill  # Apply "now" color fill to the entire row
         # Check if the label "Next" is in the current issue's labels
            if 'Next' in row[4].value:  # Assuming 'Labels' is the 5th column (index 4)
                for cell in row:
                    cell.fill = next_fill  # Apply "now" color fill to the entire row
         # Check if the label "Later" is in the current issue's labels
            if 'Later' in row[4].value:  # Assuming 'Labels' is the 5th column (index 4)
                for cell in row:
                    cell.fill = later_fill  # Apply "now" color fill to the entire row
         # Check if the label "Spike" is in the current issue's labels
            if '[SPIKE]' in row[4].value:  # Assuming 'Labels' is the 5th column (index 4)
                for cell in row:
                    cell.fill = spike_fill  # Apply "now" color fill to the entire row
                    

    # Save the workbook
    workbook.save(excel_file)
    workbook.close()

# Main execution
if __name__ == "__main__":
    issues_with_comments = get_issues_and_comments()
    if issues_with_comments:
        save_to_excel(issues_with_comments)  # Save the output to Excel
        print("Data has been saved to gitlab_issues_with_comments.xlsx")
    else:
        print("No issues found or an error occurred.")
